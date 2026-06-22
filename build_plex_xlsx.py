from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "PLEX Library"

NAVY  = "0A1628"
RED   = "C8102E"
WHITE = "FFFFFF"
LGRAY = "E8EDF3"
MGRAY = "C5CDD8"

hdr_font  = Font(name="Arial", bold=True, color=WHITE, size=9)
hdr_fill  = PatternFill("solid", fgColor=NAVY)
hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

cat_font  = Font(name="Arial", bold=True, color=WHITE, size=9)
cat_fill  = PatternFill("solid", fgColor=RED)
cat_align = Alignment(horizontal="left", vertical="center")

data_font  = Font(name="Arial", size=9)
data_fontb = Font(name="Arial", size=9, bold=True, color="0A1628")
alt_fill   = PatternFill("solid", fgColor=LGRAY)
num_align  = Alignment(horizontal="right", vertical="center")
txt_align  = Alignment(horizontal="left", vertical="center")
ctr_align  = Alignment(horizontal="center", vertical="center")

thin = Side(style="thin", color=MGRAY)
bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

headers    = ["Category","Device Name","Watts (W)","Rack Units (U)","AC Voltage","Location",
              "PoE Consumer","PoE Switch","PoE Ports","PoE Budget (W)","UPS Cap (W)","UPS Cap (VA)","RT (min)","Is UPS"]
col_widths = [28, 46, 10, 9, 9, 8, 12, 10, 9, 13, 11, 11, 8, 7]

library = [
  ("NVR / VMS Servers","Milestone Husky IVO 1800R",990,2,120,"rack",False,False,0,0,0,0,0,False),
  ("NVR / VMS Servers","Milestone Husky IVO 700R",315,2,120,"rack",False,False,0,0,0,0,0,False),
  ("NVR / VMS Servers","Milestone Husky IVO 350R",495,1,120,"rack",False,False,0,0,0,0,0,False),
  ("NVR / VMS Servers","Genetec Streamvault SV-100",58,1,120,"rack",False,False,0,0,0,0,0,False),
  ("NVR / VMS Servers","Dell PowerEdge R350",540,1,120,"rack",False,False,0,0,0,0,0,False),
  ("NVR / VMS Servers","Generic 1U Server",300,1,120,"rack",False,False,0,0,0,0,0,False),
  ("PoE Switches","Cisco CBS350 24P (185W PoE)",31,1,120,"rack",False,True,24,185,0,0,0,False),
  ("PoE Switches","Cisco CBS350 48P (370W PoE)",55,1,120,"rack",False,True,48,370,0,0,0,False),
  ("PoE Switches","Cisco CBS350 24FP (370W PoE)",42,1,120,"rack",False,True,24,370,0,0,0,False),
  ("PoE Switches","Axis D8004 PoE (60W)",2,1,120,"rack",False,True,4,60,0,0,0,False),
  ("PoE Switches","Hanwha SPOEx-200P8 (240W PoE)",55,1,120,"rack",False,True,8,240,0,0,0,False),
  ("PoE Switches","Generic Unmanaged 8P PoE (120W)",30,1,120,"rack",False,True,8,120,0,0,0,False),
  ("PoE Switches","Juniper EX4300-48P (900W PoE+)",157,1,120,"rack",False,True,48,900,0,0,0,False),
  ("PoE Switches","Juniper EX4300-48MP (1100W PoE++)",270,1,120,"rack",False,True,48,1100,0,0,0,False),
  ("Cameras (PoE)","Axis P32 Series (Indoor Dome)",8,0,120,"rack",True,False,0,0,0,0,0,False),
  ("Cameras (PoE)","Axis Q61 Series (PTZ)",60,0,120,"rack",True,False,0,0,0,0,0,False),
  ("Cameras (PoE)","Axis P14 Series (Bullet)",13,0,120,"rack",True,False,0,0,0,0,0,False),
  ("Cameras (PoE)","Hanwha QNV Series (Vandal Dome)",11,0,120,"rack",True,False,0,0,0,0,0,False),
  ("Cameras (PoE)","Hanwha XNV Series (Outdoor)",13,0,120,"rack",True,False,0,0,0,0,0,False),
  ("Cameras (PoE)","Generic IP Camera (PoE)",12,0,120,"rack",True,False,0,0,0,0,0,False),
  ("Cameras (PoE)","Generic IP PTZ (HiPoE/PoE+)",25,0,120,"rack",True,False,0,0,0,0,0,False),
  ("UPS","APC Smart-UPS 1500VA LCD (980W)",18,2,120,"rack",False,False,0,0,980,1500,6,True),
  ("UPS","APC Smart-UPS 2200VA LCD (1980W)",35,2,208,"rack",False,False,0,0,1980,2200,7,True),
  ("UPS","APC Smart-UPS 3000VA LCD (2700W)",45,2,208,"rack",False,False,0,0,2700,3000,9,True),
  ("UPS","Eaton 9PX 1500VA (1350W)",85,2,120,"rack",False,False,0,0,1350,1500,5,True),
  ("UPS","Eaton 9PX 3000VA (2700W)",170,2,208,"rack",False,False,0,0,2700,3000,5,True),
  ("Networking / Other (Rack)","Cisco C9300 24-Port (non-PoE)",64,1,120,"rack",False,False,0,0,0,0,0,False),
  ("Networking / Other (Rack)","Cisco Catalyst 9300-24S (24p SFP, non-PoE)",113,1,120,"rack",False,False,0,0,0,0,0,False),
  ("Networking / Other (Rack)","Cisco Meraki MX95 (SD-WAN / Firewall)",100,1,120,"rack",False,False,0,0,0,0,0,False),
  ("Networking / Other (Rack)","5G / LTE Cellular Gateway (1U)",30,1,120,"rack",False,False,0,0,0,0,0,False),
  ("Networking / Other (Rack)","1U Cable Management - Brush Guard",0,1,120,"rack",False,False,0,0,0,0,0,False),
  ("Networking / Other (Rack)","1U Patch Panel (passive)",0,1,120,"rack",False,False,0,0,0,0,0,False),
  ("Networking / Other (Rack)","2U Patch Panel (passive)",0,2,120,"rack",False,False,0,0,0,0,0,False),
  ("Networking / Other (Rack)","Fiber Distribution Unit (FDU 1U)",0,1,120,"rack",False,False,0,0,0,0,0,False),
  ("Networking / Other (Rack)","KVM Over IP (1U)",15,1,120,"rack",False,False,0,0,0,0,0,False),
  ("Networking / Other (Rack)","Console Server (1U)",15,1,120,"rack",False,False,0,0,0,0,0,False),
  ("Networking / Other (Rack)","PDU (Rack, metered)",10,1,120,"rack",False,False,0,0,0,0,0,False),
  ("Networking / Other (Rack)","1U Blank Panel",0,1,120,"rack",False,False,0,0,0,0,0,False),
  ("ACS Power Supplies (Wall)","LifeSafety Power FlexPower e2 (2A)",30,0,120,"wall",False,False,0,0,0,0,0,False),
  ("ACS Power Supplies (Wall)","LifeSafety Power FlexPower e4 (4A)",74,0,120,"wall",False,False,0,0,0,0,0,False),
  ("ACS Power Supplies (Wall)","LifeSafety Power FlexPower e8 (8A)",110,0,120,"wall",False,False,0,0,0,0,0,False),
  ("ACS Power Supplies (Wall)","LifeSafety Power FlexPower e12 (12A)",153,0,120,"wall",False,False,0,0,0,0,0,False),
  ("ACS Power Supplies (Wall)","Altronix AL600ULX (6A 12/24V)",100,0,120,"wall",False,False,0,0,0,0,0,False),
  ("ACS Power Supplies (Wall)","Altronix AL1012ULACM (10A 12V)",150,0,120,"wall",False,False,0,0,0,0,0,False),
  ("ACS Door Controllers (Wall)","Mercury MR-52 (2-door, 2 readers/port)",11,0,120,"wall",False,False,0,0,0,0,0,False),
  ("ACS Door Controllers (Wall)","Mercury MR-50 (1-door, 2 readers/port)",6,0,120,"wall",False,False,0,0,0,0,0,False),
  ("ACS Door Controllers (Wall)","Mercury MR-16 (16-input monitor)",3,0,120,"wall",False,False,0,0,0,0,0,False),
  ("ACS Door Controllers (Wall)","Mercury LP1502 (2-door, 2 readers/port)",12,0,120,"wall",False,False,0,0,0,0,0,False),
  ("ACS Door Controllers (Wall)","Mercury LP2500 (intelligent controller)",10,0,120,"wall",False,False,0,0,0,0,0,False),
  ("ACS Door Controllers (Wall)","Mercury LP4502 (4-door)",18,0,120,"wall",False,False,0,0,0,0,0,False),
  ("ACS Door Controllers (Wall)","HID VertX EVO V1000 (1-door)",12,0,120,"wall",False,False,0,0,0,0,0,False),
  ("ACS Door Controllers (Wall)","HID VertX EVO V2000 (2-door)",15,0,120,"wall",False,False,0,0,0,0,0,False),
  ("ACS Door Controllers (Wall)","Genetec SY-ACU222 Synergis (2-door)",14,0,120,"wall",False,False,0,0,0,0,0,False),
  ("Door Hardware / Locks (Wall)","Electric Strike, Fail Secure (12VDC)",6,0,120,"wall",False,False,0,0,0,0,0,False),
  ("Door Hardware / Locks (Wall)","Electric Strike, Fail Secure (24VDC)",6,0,120,"wall",False,False,0,0,0,0,0,False),
  ("Door Hardware / Locks (Wall)","Mortise Lock, Fail Secure (12VDC)",8,0,120,"wall",False,False,0,0,0,0,0,False),
  ("Door Hardware / Locks (Wall)","Mortise Lock, Fail Secure (24VDC)",8,0,120,"wall",False,False,0,0,0,0,0,False),
  ("Door Hardware / Locks (Wall)","Schlage L9092EU Electrified Mortise Lockset (24VDC)",8,0,120,"wall",False,False,0,0,0,0,0,False),
  ("Door Hardware / Locks (Wall)","Electromagnetic Lock, Fail Safe",5,0,120,"wall",False,False,0,0,0,0,0,False),
  ("Door Hardware / Locks (Wall)","Electric Latch Retraction (ELR)",18,0,120,"wall",False,False,0,0,0,0,0,False),
  ("Door Hardware / Locks (Wall)","Request-to-Exit (REX) Sensor",2,0,120,"wall",False,False,0,0,0,0,0,False),
  ("Field / Wall Equipment","DAS BDA Head End Unit",100,0,120,"wall",False,False,0,0,0,0,0,False),
  ("Field / Wall Equipment","EERC Master Station",25,0,120,"wall",False,False,0,0,0,0,0,False),
  ("Field / Wall Equipment","EERC Remote Station",15,0,120,"wall",False,False,0,0,0,0,0,False),
  ("Field / Wall Equipment","Intercom Master Station (wall)",20,0,120,"wall",False,False,0,0,0,0,0,False),
  ("Field / Wall Equipment","Generic Wall-Mount Enclosure",30,0,120,"wall",False,False,0,0,0,0,0,False),
]

# Title row
ws.merge_cells("A1:N1")
tc = ws["A1"]
tc.value = "PLEX v2.32 — Device Power Library   |   Limited Energy eXperts   |   limitedenergy.net/plex"
tc.font  = Font(name="Arial", bold=True, color=WHITE, size=10)
tc.fill  = PatternFill("solid", fgColor=NAVY)
tc.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 22

# Subtitle / notes row
ws.merge_cells("A2:N2")
sc = ws["A2"]
sc.value = ("Watts (W) = 90% of nameplate max power consumption.  "
            "PoE switch W = chassis overhead only; PoE budget tracked separately.  "
            "UPS W = self-consumption (losses).  "
            "Board W includes reader port loads.")
sc.font  = Font(name="Arial", italic=True, color="444444", size=8)
sc.fill  = PatternFill("solid", fgColor="D0D8E4")
sc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[2].height = 28

# Header row
for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=3, column=col, value=h)
    cell.font      = hdr_font
    cell.fill      = hdr_fill
    cell.alignment = hdr_align
    cell.border    = bdr
    ws.column_dimensions[get_column_letter(col)].width = w
ws.row_dimensions[3].height = 32

# Data rows with category banners
current_cat = None
excel_row = 4

for row in library:
    cat = row[0]

    if cat != current_cat:
        ws.merge_cells(start_row=excel_row, start_column=1,
                       end_row=excel_row, end_column=14)
        cc = ws.cell(row=excel_row, column=1, value=cat)
        cc.font      = cat_font
        cc.fill      = cat_fill
        cc.alignment = cat_align
        cc.border    = bdr
        ws.row_dimensions[excel_row].height = 16
        excel_row += 1
        current_cat = cat

    is_alt = (excel_row % 2 == 0)
    row_fill = PatternFill("solid", fgColor=LGRAY) if is_alt else None

    for col, val in enumerate(row, start=1):
        cell = ws.cell(row=excel_row, column=col)
        if isinstance(val, bool):
            cell.value     = "Y" if val else ""
            cell.font      = data_font
            cell.alignment = ctr_align
        elif col == 2:
            cell.value     = val
            cell.font      = data_font
            cell.alignment = txt_align
        elif col == 1:
            cell.value     = val
            cell.font      = data_font
            cell.alignment = txt_align
        elif col in (3, 4, 5, 9, 10, 11, 12, 13):
            cell.value     = val if val != 0 else None
            cell.alignment = num_align
            cell.font      = data_fontb if col == 3 and val > 0 else data_font
        elif col == 6:
            cell.value     = val.capitalize()
            cell.font      = data_font
            cell.alignment = ctr_align
        else:
            cell.value     = val
            cell.font      = data_font
            cell.alignment = txt_align
        if row_fill:
            cell.fill = row_fill
        cell.border = bdr

    ws.row_dimensions[excel_row].height = 15
    excel_row += 1

ws.freeze_panes = "A4"
ws.auto_filter.ref = "A3:N3"

out = r"D:\DEV\LEX\PLEX_Library_v2.32.xlsx"
wb.save(out)
print(f"Saved: {out}")
